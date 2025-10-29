# dao/user_dao_excel.py

import openpyxl
from models.user import User
from dao.user_dao_base import UserDAOBase
import os

class UserDAOExcel(UserDAOBase):
    def __init__(self, filepath: str):
        self.filepath = filepath
        # If the file doesn't exist, create it with a header row
        if not os.path.exists(filepath):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["id", "name", "email"])
            workbook.save(filepath)

    def add_user(self, user: User):
        # Load the existing workbook and get the active sheet
        workbook = openpyxl.load_workbook(self.filepath)
        sheet = workbook.active

        # Determine the new user ID based on the number of rows
        # Subtract 1 because the first row is the header
        new_id = sheet.max_row
        
        # We need to call get_all_users to get the correct next ID based on the content
        # An Excel file might have empty rows, max_row is not reliable for ID
        # A more robust way is to read all and find max ID, but for simplicity
        # we'll use a count of non-header rows or the max row.
        # Let's use a count of non-header rows for simplicity in this structure
        user_id = len(self.get_all_users()) + 1

        # Append the new user data as a row
        sheet.append([user_id, user.name, user.email])
        
        # Save the changes
        workbook.save(self.filepath)

    def get_all_users(self) -> list[User]:
        users = []
        try:
            workbook = openpyxl.load_workbook(self.filepath)
            sheet = workbook.active
            
            # Iterate through rows, skipping the header (row 1)
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                # Each row is a tuple: (id, name, email)
                id_, name, email = row
                users.append(User(int(id_), name, email))
        except FileNotFoundError:
            # If the file is somehow deleted between init and call, return empty list
            pass
        return users